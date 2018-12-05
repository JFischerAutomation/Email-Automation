from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
import os
import datetime
import sys
import smtplib
wb = openpyxl.load_workbook('Deposit Reminder.xlsx')
ws = wb['Reminder']
date, name, l_Name, email, emailReminder = [], [], [], [], []


#reads the spreadsheet to populate the GUI
for i in range(2,ws.max_row+1):
    if ws.cell(row = i, column = 1).value != None:
        date.append(ws.cell(row = i,column = 3).value)
        name.append(ws.cell(row = i,column =1).value)
        l_Name.append(ws.cell(row = i,column =2).value)
        email.append(ws.cell(row =i, column = 4).value)
        emailReminder.append(ws.cell(row = i,column = 5).value)
#looks at dates that are past due and 7 days ahead to populate the GUI


reference = datetime.datetime.now() + datetime.timedelta(days = 7)
reference = reference.strftime('%m-%d-%y')
pastDueName, pastDueLname, pastDueDate, pastDueEmail = [], [], [], []
emailDictionary = dict(zip(email,emailReminder))
date = [d.strftime('%m-%d-%y')for d in date]

for i in range(len(date)):
    if date[i] <= reference and emailReminder[i] == ('No' or 'no')  and emailReminder!='Yes':
        pastDueName.append(name[i])
        pastDueLname.append(l_Name[i])
        pastDueDate.append(date[i])
        pastDueEmail.append(email[i])
max_row = len(pastDueName)
myEmail = #email address to be emailed from
myEmailPassword = # password for email login
smtpObj = smtplib.SMTP('smtp.gmail.com',587)
smtpObj.starttls()
smtpObj.login(myEmail,password)



#custom class created to allow iteration over buttons created
class PushButts(QtWidgets.QPushButton):

    def setIndex(self,x):
        self.index = x

    def getIndex(self):
        return self.index

    def setF(self, f):
        self.f = f

    def getF(self):
        return self.f

    def applyF(self):
        f = self.getF()
        f(self.getIndex())

#Main GUI window
class Ui_MainWindow(object):

    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(786, 608)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(0, 0, 800, 1000))
        self.tableWidget.setColumnCount(4)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setRowCount(max_row)

        #when this button is clicked, will send an automated email message to the email's reminding them of their deposits
        self.pushButton = QtWidgets.QPushButton('Email Button',self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(650,30,100,25))
        self.pushButton.clicked.connect(self.saveFile)
        self.pushButton.clicked.connect(self.printEmail)


        #creates number of rows based upon criteria established for past due deposits
        for i in range(max_row):
            self.tableWidget.setItem(i,0,QtWidgets.QTableWidgetItem(pastDueName[i]))
            self.tableWidget.setItem(i,1,QtWidgets.QTableWidgetItem(pastDueLname[i]))
            self.tableWidget.setItem(i,2,QtWidgets.QTableWidgetItem(pastDueDate[i]))
            self.tableWidget.setItem(i,3,QtWidgets.QTableWidgetItem(pastDueEmail[i]))
            header = self.tableWidget.horizontalHeader()
            header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)



        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 786, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        #function that sends the emails to the one's that fit the criteria established
    def printEmail(self, index):
        for i in range(max_row):
            message = 'Hello! This is an automated message reminding you of your deposit that is due on '+str(pastDueDate[i])
            smtpObj.sendmail(myEmail,pastDueEmail[i],message)
        
        #function that modifies the excel file so that when the email goes out, it tracks it and then disallows it from populating the GUI next time it is opened
    def saveFile(self,index):
        for i in range(max_row):
            emailDictionary[pastDueEmail[i]]='Yes'
            emailReminder= list(emailDictionary.values())
        for x in range(len(emailReminder)):
            ws.cell(row=x+2,column = 5).value = emailReminder[x]
            wb.save('Deposit Reminder.xlsx')





if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

smtpObj.quit()